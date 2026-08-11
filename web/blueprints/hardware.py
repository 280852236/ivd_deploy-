#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""IVD平台 - HARDWARE模块"""

from flask import Blueprint, request, jsonify, session, redirect, url_for, make_response, render_template_string
from psycopg2.extras import RealDictCursor, execute_batch
import shared
from shared import api_login_required, api_super_admin_required, login_required
from services.cache import api_cache, invalidate_cache
import json
import logging

logger = logging.getLogger(__name__)

def _hardware_failure_table(model):
    return shared.resolve_table(model, 'hardware_failures')

def _pcba_compat_table(model):
    from services.data_init import ensure_pcba_compat_table, get_pcba_compat_table
    ensure_pcba_compat_table(model)
    shared._table_cache.pop(get_pcba_compat_table(model), None)
    return shared.resolve_table(model, 'pcba_compat')

def _bootloader_compat_table(model):
    from services.data_init import ensure_bootloader_compat_table, get_bootloader_compat_table
    ensure_bootloader_compat_table(model)
    shared._table_cache.pop(get_bootloader_compat_table(model), None)
    return shared.resolve_table(model, 'bootloader_compat')

hardware_bp = Blueprint('hardware', __name__)


@hardware_bp.route('/api/hardware-failures', methods=['GET'])
@api_cache(ttl=60, key_prefix='hw')
def get_hardware_failures():
    model = request.args.get('model', '').strip()
    page = max(1, request.args.get('page', 1, type=int))
    per_page = min(100, max(1, request.args.get('per_page', 20, type=int)))
    if not model:
        return jsonify({'results': [], 'total': 0, 'page': page, 'per_page': per_page})
    tbl = _hardware_failure_table(model)
    if not tbl:
        return jsonify({'results': [], 'total': 0, 'page': page, 'per_page': per_page})
    img_tbl = shared.safe_img_table(model, 'hardware_failure_images')
    with shared.db_connection() as conn:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(f'SELECT COUNT(*) AS total FROM {tbl} h')
        total = cur.fetchone()['total']
        cur.execute(f'SELECT h.id, %s AS model, h.phenomenon, h.cause, h.workaround, h.process, h.suggestion, h.solution, h.created_at, h.updated_at, COALESCE(img_cnt.cnt, 0) AS image_count FROM {tbl} h LEFT JOIN (SELECT failure_id, COUNT(*) AS cnt FROM {img_tbl} GROUP BY failure_id) img_cnt ON img_cnt.failure_id = h.id ORDER BY h.created_at DESC LIMIT %s OFFSET %s', [model, per_page, (page - 1) * per_page])
        rows = cur.fetchall()
        for row in rows:
            shared.format_row_timestamps(row)
        return jsonify({'results': [dict(r) for r in rows], 'total': total, 'page': page, 'per_page': per_page})



@hardware_bp.route('/api/hardware-failures', methods=['POST'])
@api_login_required
def add_hardware_failure():
    model = request.form.get('model', '').strip()
    phenomenon = request.form.get('phenomenon', '').strip()
    cause = request.form.get('cause', '').strip()
    workaround = request.form.get('workaround', '').strip()
    process = request.form.get('process', '').strip()
    suggestion = request.form.get('suggestion', '').strip()
    solution = request.form.get('solution', '').strip()
    if not model or not phenomenon:
        return jsonify({'error': '型号和故障现象为必填项'}), 400
    tbl = _hardware_failure_table(model)
    if not tbl:
        return jsonify({'error': f'型号 {model} 的硬件故障表不存在'}), 404
    img_tbl = shared.safe_img_table(model, 'hardware_failure_images')
    try:
        with shared.db_connection() as conn:
            cur = conn.cursor(cursor_factory=RealDictCursor)
            cur.execute(f'INSERT INTO {tbl} (phenomenon, cause, workaround, process, suggestion, solution) VALUES (%s, %s, %s, %s, %s, %s) RETURNING id', (phenomenon, cause, workaround, process, suggestion, solution))
            failure_id = cur.fetchone()['id']
            for file in request.files.getlist('images'):
                if file and file.content_type in shared.ALLOWED_IMAGE_TYPES:
                    img_data = file.read()
                    cur.execute(f'INSERT INTO {img_tbl} (failure_id, image_data) VALUES (%s, %s)', (failure_id, img_data))
            conn.commit()
        return jsonify({'success': True, 'id': failure_id})
    except Exception as e:
        logger.exception('添加硬件故障失败')
        return jsonify({'error': '保存失败，请联系管理员'}), 500



@hardware_bp.route('/api/hardware-failures/<model>/<int:failure_id>', methods=['PUT'])
@api_login_required
def update_hardware_failure(model, failure_id):
    phenomenon = request.form.get('phenomenon', '').strip()
    cause = request.form.get('cause', '').strip()
    workaround = request.form.get('workaround', '').strip()
    process = request.form.get('process', '').strip()
    suggestion = request.form.get('suggestion', '').strip()
    solution = request.form.get('solution', '').strip()
    if not phenomenon:
        return jsonify({'error': '故障现象为必填项'}), 400
    tbl = _hardware_failure_table(model)
    if not tbl:
        return jsonify({'error': f'型号 {model} 的硬件故障表不存在'}), 404
    img_tbl = shared.safe_img_table(model, 'hardware_failure_images')
    try:
        with shared.db_connection() as conn:
            cur = conn.cursor(cursor_factory=RealDictCursor)
            cur.execute(f'UPDATE {tbl} SET phenomenon=%s, cause=%s, workaround=%s, process=%s, suggestion=%s, solution=%s, updated_at=NOW() WHERE id=%s', (phenomenon, cause, workaround, process, suggestion, solution, failure_id))
            for file in request.files.getlist('images'):
                if file and file.content_type in shared.ALLOWED_IMAGE_TYPES:
                    img_data = file.read()
                    cur.execute(f'INSERT INTO {img_tbl} (failure_id, image_data) VALUES (%s, %s)', (failure_id, img_data))
            conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        logger.exception('更新硬件故障失败')
        return jsonify({'error': '更新失败，请联系管理员'}), 500



@hardware_bp.route('/api/hardware-failures/<model>/<int:failure_id>', methods=['DELETE'])
@api_login_required
def delete_hardware_failure(model, failure_id):
    tbl = _hardware_failure_table(model)
    if not tbl:
        return jsonify({'error': f'型号 {model} 的硬件故障表不存在'}), 404
    with shared.db_connection() as conn:
        cur = conn.cursor()
        cur.execute(f'DELETE FROM {tbl} WHERE id = %s', (failure_id,))
        conn.commit()
    shared.audit_log('delete_hardware_failure', target_type='hardware_failure', target_id=failure_id, detail=f'删除 {model} 硬件故障#{failure_id}')
    return jsonify({'success': True})



@hardware_bp.route('/api/hardware-failures/<model>/<int:failure_id>/images', methods=['GET'])
def get_hardware_failure_images(model, failure_id):
    img_tbl = shared.safe_img_table(model, 'hardware_failure_images')
    with shared.db_connection() as conn:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(f'SELECT id FROM {img_tbl} WHERE failure_id = %s ORDER BY id', (failure_id,))
        rows = cur.fetchall()
    return jsonify([{'id': r['id']} for r in rows])



@hardware_bp.route('/api/hardware-failures/<model>/<int:failure_id>/images/<int:image_id>', methods=['GET'])
def get_hardware_failure_image(model, failure_id, image_id):
    from flask import Response
    img_tbl = shared.safe_img_table(model, 'hardware_failure_images')
    with shared.db_connection() as conn:
        cur = conn.cursor()
        cur.execute(f'SELECT image_data, image_mime FROM {img_tbl} WHERE id = %s AND failure_id = %s', (image_id, failure_id))
        row = cur.fetchone()
        if not row or not row[0]:
            return 'Image not found', 404
        img_data, img_mime = row[0], row[1] or 'image/jpeg'
    response = Response(img_data, mimetype=img_mime)
    response.headers['Cache-Control'] = 'public, max-age=604800'
    response.headers['Content-Length'] = len(img_data)
    return response



@hardware_bp.route('/api/hardware-failures/<model>/<int:failure_id>/images/<int:image_id>', methods=['DELETE'])
@api_login_required
def delete_hardware_failure_image(model, failure_id, image_id):
    img_tbl = shared.safe_img_table(model, 'hardware_failure_images')
    with shared.db_connection() as conn:
        cur = conn.cursor()
        cur.execute(f'DELETE FROM {img_tbl} WHERE id = %s AND failure_id = %s', (image_id, failure_id))
        conn.commit()
    return jsonify({'success': True})



@hardware_bp.route('/api/hardware-failures/search', methods=['GET'])
def search_hardware_failures():
    q = request.args.get('q', '').strip()
    model = request.args.get('model', '').strip()
    page = max(1, request.args.get('page', 1, type=int))
    per_page = min(100, max(1, request.args.get('per_page', 20, type=int)))
    if not q:
        return jsonify({'results': [], 'total': 0, 'page': page, 'per_page': per_page})
    if not model:
        return jsonify({'results': [], 'total': 0, 'page': page, 'per_page': per_page})
    tbl = _hardware_failure_table(model)
    if not tbl:
        return jsonify({'results': [], 'total': 0, 'page': page, 'per_page': per_page})
    img_tbl = shared.safe_img_table(model, 'hardware_failure_images')
    with shared.db_connection() as conn:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        search_pattern = f'%{q}%'
        where_clause = "h.phenomenon ILIKE %s OR h.cause ILIKE %s OR h.workaround ILIKE %s OR h.process ILIKE %s OR h.suggestion ILIKE %s OR h.solution ILIKE %s"
        cur.execute(f'SELECT COUNT(*) AS total FROM {tbl} h WHERE {where_clause}', [search_pattern]*6)
        total = cur.fetchone()['total']
        cur.execute(f"""
            SELECT * FROM (
                SELECT h.id, %s AS model, h.phenomenon, h.cause, h.workaround,
                       h.process, h.suggestion, h.solution,
                       h.created_at, h.updated_at,
                       COALESCE(img_cnt.cnt, 0) AS image_count
                 FROM {tbl} h
                 LEFT JOIN (SELECT failure_id, COUNT(*) AS cnt FROM {img_tbl} GROUP BY failure_id) img_cnt ON img_cnt.failure_id = h.id
                WHERE {where_clause}
                ORDER BY h.created_at DESC
            ) sub LIMIT %s OFFSET %s
        """, [model, search_pattern, search_pattern, search_pattern, search_pattern, search_pattern, search_pattern, per_page, (page - 1) * per_page])
        rows = cur.fetchall()
        for row in rows:
            shared.format_row_timestamps(row)
        return jsonify({'results': [dict(r) for r in rows], 'total': total, 'page': page, 'per_page': per_page})


@hardware_bp.route('/hardware-failures')
@login_required
def hardware_failures_page():
    series = request.args.get('series', 'SMART').upper()
    model = request.args.get('model', '')
    if series not in ('SMART', 'VENUS'):
        series = 'SMART'
    is_admin = session.get('admin_logged_in', False)
    return render_template_string(shared.get_template('HARDWARE_FAILURES_HTML'), series=series, model=model, is_admin=is_admin)


# ========== 电路板兼容表 API ==========

_PCBA_FIELDS = ['pcba_code', 'pcb_code', 'pcb_silkscreen', 'latest_version', 'board_name', 'special_note', 'pcba_version_compat', 'compat_description']
_BOOTLOADER_FIELDS = ['board_mnemonic', 'board_name', 'bootloader_version', 'bootloader_compat_note', 'no_bootloader_version', 'no_bootloader_compat_note']

_PCBA_HEADER_MAP = {
    'pcba_code': ['PCBA编码', 'PCBA Code', 'pcba_code', 'PCBA编号'],
    'pcb_code': ['PCB编码', 'PCB Code', 'pcb_code', 'PCB编号'],
    'pcb_silkscreen': ['PCB丝印', 'PCB Silkscreen', 'pcb_silkscreen'],
    'latest_version': ['最新版本', 'Latest Version', 'latest_version'],
    'board_name': ['电路板名称', 'Board Name', 'board_name', '板卡名称'],
    'special_note': ['特殊说明', 'Special Note', 'special_note'],
    'pcba_version_compat': ['PCBA版本兼容性', 'PCBA Version Compat', 'pcba_version_compat'],
    'compat_description': ['兼容性说明', 'Compat Description', 'compat_description'],
}

_BOOTLOADER_HEADER_MAP = {
    'board_mnemonic': ['板卡助记码', 'Board Mnemonic', 'board_mnemonic', '助记码'],
    'board_name': ['电路板名称', 'Board Name', 'board_name', '板卡名称'],
    'bootloader_version': ['底层版本（Bootloader）', 'Bootloader Version', 'bootloader_version', '底层版本(Bootloader)'],
    'bootloader_compat_note': ['底层版本（Bootloader）兼容性说明', 'Bootloader Compat Note', 'bootloader_compat_note', '底层版本(Bootloader)兼容性说明'],
    'no_bootloader_version': ['底层版本（无Bootloader）', 'No Bootloader Version', 'no_bootloader_version', '底层版本(无Bootloader)'],
    'no_bootloader_compat_note': ['底层版本（无Bootloader）兼容性说明', 'No Bootloader Compat Note', 'no_bootloader_compat_note', '底层版本(无Bootloader)兼容性说明'],
}


@hardware_bp.route('/api/board-compat/pcba', methods=['GET'])
@api_cache(ttl=300, key_prefix='pcba')
def get_pcba_compat():
    model = request.args.get('model', '').strip()
    keyword = request.args.get('keyword', '').strip()
    page = max(1, request.args.get('page', 1, type=int))
    per_page = min(200, max(1, request.args.get('per_page', 100, type=int)))
    if not model:
        return jsonify({'results': [], 'total': 0, 'page': page, 'per_page': per_page})
    tbl = _pcba_compat_table(model)
    if not tbl:
        return jsonify({'results': [], 'total': 0, 'page': page, 'per_page': per_page})
    offset = (page - 1) * per_page
    with shared.db_connection() as conn:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        if keyword:
            search_pattern = f'%{keyword}%'
            conditions = ' OR '.join(f'{f} ILIKE %s' for f in _PCBA_FIELDS)
            cur.execute(f'SELECT COUNT(*) AS total FROM {tbl} WHERE {conditions}', [search_pattern]*len(_PCBA_FIELDS))
            total = cur.fetchone()['total']
            cur.execute(f'SELECT * FROM {tbl} WHERE {conditions} ORDER BY id LIMIT %s OFFSET %s', [search_pattern]*len(_PCBA_FIELDS) + [per_page, offset])
        else:
            cur.execute(f'SELECT COUNT(*) AS total FROM {tbl}')
            total = cur.fetchone()['total']
            cur.execute(f'SELECT * FROM {tbl} ORDER BY id LIMIT %s OFFSET %s', [per_page, offset])
        rows = cur.fetchall()
        for row in rows:
            shared.format_row_timestamps(row)
        return jsonify({'results': [dict(r) for r in rows], 'total': total, 'page': page, 'per_page': per_page})


@hardware_bp.route('/api/board-compat/bootloader', methods=['GET'])
@api_cache(ttl=300, key_prefix='boot')
def get_bootloader_compat():
    model = request.args.get('model', '').strip()
    keyword = request.args.get('keyword', '').strip()
    page = max(1, request.args.get('page', 1, type=int))
    per_page = min(200, max(1, request.args.get('per_page', 100, type=int)))
    if not model:
        return jsonify({'results': [], 'total': 0, 'page': page, 'per_page': per_page})
    tbl = _bootloader_compat_table(model)
    if not tbl:
        return jsonify({'results': [], 'total': 0, 'page': page, 'per_page': per_page})
    offset = (page - 1) * per_page
    with shared.db_connection() as conn:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        if keyword:
            search_pattern = f'%{keyword}%'
            conditions = ' OR '.join(f'{f} ILIKE %s' for f in _BOOTLOADER_FIELDS)
            cur.execute(f'SELECT COUNT(*) AS total FROM {tbl} WHERE {conditions}', [search_pattern]*len(_BOOTLOADER_FIELDS))
            total = cur.fetchone()['total']
            cur.execute(f'SELECT * FROM {tbl} WHERE {conditions} ORDER BY id LIMIT %s OFFSET %s', [search_pattern]*len(_BOOTLOADER_FIELDS) + [per_page, offset])
        else:
            cur.execute(f'SELECT COUNT(*) AS total FROM {tbl}')
            total = cur.fetchone()['total']
            cur.execute(f'SELECT * FROM {tbl} ORDER BY id LIMIT %s OFFSET %s', [per_page, offset])
        rows = cur.fetchall()
        for row in rows:
            shared.format_row_timestamps(row)
        return jsonify({'results': [dict(r) for r in rows], 'total': total, 'page': page, 'per_page': per_page})


@hardware_bp.route('/api/board-compat/pcba', methods=['POST'])
@api_login_required
def add_pcba_compat():
    if not session.get('admin_logged_in'):
        return jsonify({'error': '请先登录管理员'}), 401
    data = request.get_json(force=True)
    model = data.get('model', '').strip()
    pcba_code = data.get('pcba_code', '').strip()
    if not model or not pcba_code:
        return jsonify({'error': '型号和PCBA编码为必填项'}), 400
    tbl = _pcba_compat_table(model)
    if not tbl:
        return jsonify({'error': f'型号 {model} 的PCBA兼容表不存在'}), 404
    vals = {f: data.get(f, '').strip() for f in _PCBA_FIELDS}
    if not vals['pcba_code']:
        return jsonify({'error': 'PCBA编码为必填项'}), 400
    try:
        with shared.db_connection() as conn:
            cur = conn.cursor(cursor_factory=RealDictCursor)
            cols = ', '.join(_PCBA_FIELDS)
            placeholders = ', '.join(['%s']*len(_PCBA_FIELDS))
            cur.execute(f'INSERT INTO {tbl} ({cols}) VALUES ({placeholders}) RETURNING id', [vals[f] for f in _PCBA_FIELDS])
            new_id = cur.fetchone()['id']
            conn.commit()
        return jsonify({'success': True, 'id': new_id})
    except Exception as e:
        if 'unique' in str(e).lower() or 'duplicate' in str(e).lower():
            return jsonify({'error': f'PCBA编码 {pcba_code} 已存在'}), 409
        logger.exception('添加PCBA兼容记录失败')
        return jsonify({'error': '添加失败，请联系管理员'}), 500


@hardware_bp.route('/api/board-compat/pcba/<model>/<int:row_id>', methods=['PUT'])
@api_login_required
def update_pcba_compat(model, row_id):
    if not session.get('admin_logged_in'):
        return jsonify({'error': '请先登录管理员'}), 401
    data = request.get_json(force=True)
    tbl = _pcba_compat_table(model)
    if not tbl:
        return jsonify({'error': f'型号 {model} 的PCBA兼容表不存在'}), 404
    vals = {f: data.get(f, '').strip() for f in _PCBA_FIELDS}
    try:
        with shared.db_connection() as conn:
            cur = conn.cursor()
            set_clause = ', '.join(f'{f}=%s' for f in _PCBA_FIELDS) + ', updated_at=NOW()'
            cur.execute(f'UPDATE {tbl} SET {set_clause} WHERE id=%s', [vals[f] for f in _PCBA_FIELDS] + [row_id])
            conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        logger.exception('更新PCBA兼容记录失败')
        return jsonify({'error': '更新失败，请联系管理员'}), 500


@hardware_bp.route('/api/board-compat/pcba/<model>/<int:row_id>', methods=['DELETE'])
@api_login_required
def delete_pcba_compat(model, row_id):
    if not session.get('admin_logged_in'):
        return jsonify({'error': '请先登录管理员'}), 401
    tbl = _pcba_compat_table(model)
    if not tbl:
        return jsonify({'error': f'型号 {model} 的PCBA兼容表不存在'}), 404
    with shared.db_connection() as conn:
        cur = conn.cursor()
        cur.execute(f'DELETE FROM {tbl} WHERE id = %s', (row_id,))
        conn.commit()
    return jsonify({'success': True})


@hardware_bp.route('/api/board-compat/bootloader', methods=['POST'])
@api_login_required
def add_bootloader_compat():
    if not session.get('admin_logged_in'):
        return jsonify({'error': '请先登录管理员'}), 401
    data = request.get_json(force=True)
    model = data.get('model', '').strip()
    board_mnemonic = data.get('board_mnemonic', '').strip()
    if not model or not board_mnemonic:
        return jsonify({'error': '型号和板卡助记码为必填项'}), 400
    tbl = _bootloader_compat_table(model)
    if not tbl:
        return jsonify({'error': f'型号 {model} 的底层兼容表不存在'}), 404
    vals = {f: data.get(f, '').strip() for f in _BOOTLOADER_FIELDS}
    if not vals['board_mnemonic']:
        return jsonify({'error': '板卡助记码为必填项'}), 400
    try:
        with shared.db_connection() as conn:
            cur = conn.cursor(cursor_factory=RealDictCursor)
            cols = ', '.join(_BOOTLOADER_FIELDS)
            placeholders = ', '.join(['%s']*len(_BOOTLOADER_FIELDS))
            cur.execute(f'INSERT INTO {tbl} ({cols}) VALUES ({placeholders}) RETURNING id', [vals[f] for f in _BOOTLOADER_FIELDS])
            new_id = cur.fetchone()['id']
            conn.commit()
        return jsonify({'success': True, 'id': new_id})
    except Exception as e:
        if 'unique' in str(e).lower() or 'duplicate' in str(e).lower():
            return jsonify({'error': f'板卡助记码 {board_mnemonic} 已存在'}), 409
        logger.exception('添加PCBA兼容记录失败')
        return jsonify({'error': '添加失败，请联系管理员'}), 500


@hardware_bp.route('/api/board-compat/bootloader/<model>/<int:row_id>', methods=['PUT'])
@api_login_required
def update_bootloader_compat(model, row_id):
    if not session.get('admin_logged_in'):
        return jsonify({'error': '请先登录管理员'}), 401
    data = request.get_json(force=True)
    tbl = _bootloader_compat_table(model)
    if not tbl:
        return jsonify({'error': f'型号 {model} 的底层兼容表不存在'}), 404
    vals = {f: data.get(f, '').strip() for f in _BOOTLOADER_FIELDS}
    try:
        with shared.db_connection() as conn:
            cur = conn.cursor()
            set_clause = ', '.join(f'{f}=%s' for f in _BOOTLOADER_FIELDS) + ', updated_at=NOW()'
            cur.execute(f'UPDATE {tbl} SET {set_clause} WHERE id=%s', [vals[f] for f in _BOOTLOADER_FIELDS] + [row_id])
            conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        logger.exception('更新底层兼容记录失败')
        return jsonify({'error': '更新失败，请联系管理员'}), 500


@hardware_bp.route('/api/board-compat/bootloader/<model>/<int:row_id>', methods=['DELETE'])
@api_login_required
def delete_bootloader_compat(model, row_id):
    if not session.get('admin_logged_in'):
        return jsonify({'error': '请先登录管理员'}), 401
    tbl = _bootloader_compat_table(model)
    if not tbl:
        return jsonify({'error': f'型号 {model} 的底层兼容表不存在'}), 404
    with shared.db_connection() as conn:
        cur = conn.cursor()
        cur.execute(f'DELETE FROM {tbl} WHERE id = %s', (row_id,))
        conn.commit()
    return jsonify({'success': True})


@hardware_bp.route('/api/board-compat/pcba/import', methods=['POST'])
@api_login_required
def import_pcba_compat():
    if not session.get('admin_logged_in'):
        return jsonify({'error': '请先登录管理员'}), 401
    model = request.form.get('model', '').strip()
    if not model:
        return jsonify({'error': '请指定型号'}), 400
    if 'file' not in request.files:
        return jsonify({'error': '请上传文件'}), 400
    file = request.files['file']
    if not file.filename:
        return jsonify({'error': '请上传文件'}), 400
    tbl = _pcba_compat_table(model)
    if not tbl:
        return jsonify({'error': f'型号 {model} 的PCBA兼容表不存在'}), 404
    try:
        rows_data = _parse_pcba_import_file(file)
    except Exception as e:
        return jsonify({'error': '文件解析失败'}), 400
    added = 0
    skipped = 0
    errors = []
    batch_data = []
    for i, row in enumerate(rows_data):
        if not row.get('pcba_code'):
            skipped += 1
            continue
        batch_data.append([row.get(f, '') for f in _PCBA_FIELDS])
    if batch_data:
        try:
            with shared.db_connection() as conn:
                cur = conn.cursor()
                cols = ', '.join(_PCBA_FIELDS)
                placeholders = ', '.join(['%s']*len(_PCBA_FIELDS))
                update_cols = ", ".join(f"{f}=EXCLUDED.{f}" for f in _PCBA_FIELDS if f != "pcba_code")
                sql = f'INSERT INTO {tbl} ({cols}) VALUES ({placeholders}) ON CONFLICT (pcba_code) DO UPDATE SET {update_cols}, updated_at=NOW()'
                execute_batch(cur, sql, batch_data, page_size=100)
                added = len(batch_data)
                conn.commit()
        except Exception as e:
            errors.append('批量插入失败')
            skipped += len(batch_data)
    return jsonify({'success': True, 'added': added, 'skipped': skipped, 'errors': errors[:10]})


@hardware_bp.route('/api/board-compat/bootloader/import', methods=['POST'])
@api_login_required
def import_bootloader_compat():
    if not session.get('admin_logged_in'):
        return jsonify({'error': '请先登录管理员'}), 401
    model = request.form.get('model', '').strip()
    if not model:
        return jsonify({'error': '请指定型号'}), 400
    if 'file' not in request.files:
        return jsonify({'error': '请上传文件'}), 400
    file = request.files['file']
    if not file.filename:
        return jsonify({'error': '请上传文件'}), 400
    tbl = _bootloader_compat_table(model)
    if not tbl:
        return jsonify({'error': f'型号 {model} 的底层兼容表不存在'}), 404
    try:
        rows_data = _parse_import_file(file, _BOOTLOADER_HEADER_MAP, _BOOTLOADER_FIELDS)
    except Exception as e:
        return jsonify({'error': '文件解析失败'}), 400
    added = 0
    skipped = 0
    errors = []
    batch_data = []
    for i, row in enumerate(rows_data):
        if not row.get('board_mnemonic'):
            skipped += 1
            continue
        batch_data.append([row.get(f, '') for f in _BOOTLOADER_FIELDS])
    if batch_data:
        try:
            with shared.db_connection() as conn:
                cur = conn.cursor()
                cols = ', '.join(_BOOTLOADER_FIELDS)
                placeholders = ', '.join(['%s']*len(_BOOTLOADER_FIELDS))
                update_cols = ", ".join(f"{f}=EXCLUDED.{f}" for f in _BOOTLOADER_FIELDS if f != "board_mnemonic")
                sql = f'INSERT INTO {tbl} ({cols}) VALUES ({placeholders}) ON CONFLICT (board_mnemonic) DO UPDATE SET {update_cols}, updated_at=NOW()'
                execute_batch(cur, sql, batch_data, page_size=100)
                added = len(batch_data)
                conn.commit()
        except Exception as e:
            errors.append('批量插入失败')
            skipped += len(batch_data)
    return jsonify({'success': True, 'added': added, 'skipped': skipped, 'errors': errors[:10]})


def _parse_import_file(file, header_map, fields):
    filename = file.filename.lower()
    if filename.endswith('.csv'):
        import csv
        import io
        stream = io.StringIO(file.read().decode('utf-8-sig'), newline=None)
        reader = csv.DictReader(stream)
        raw_rows = list(reader)
    elif filename.endswith(('.xlsx', '.xls')):
        try:
            import openpyxl
        except ImportError:
            raise ValueError('请安装openpyxl: pip install openpyxl')
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
        headers = [str(cell.value or '').strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        raw_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = {}
            for i in range(min(len(headers), len(row))):
                val = row[i]
                if val is None:
                    val = ''
                elif isinstance(val, str):
                    val = val.strip()
                else:
                    val = str(val).strip()
                col_key = headers[i] if headers[i] else f'_col_{i}'
                if col_key in row_data:
                    row_data[col_key] += '\n' + val
                else:
                    row_data[col_key] = val
            raw_rows.append(row_data)
        wb.close()
    else:
        raise ValueError('仅支持 .xlsx、.xls、.csv 格式')
    reverse_map = {}
    for db_field, aliases in header_map.items():
        for alias in aliases:
            reverse_map[alias.strip().lower()] = db_field
    result = []
    for raw_row in raw_rows:
        mapped = {}
        for col_name, value in raw_row.items():
            val = str(value or '').strip()
            if val == '/':
                val = ''
            db_field = reverse_map.get(col_name.strip().lower())
            if db_field:
                if db_field in mapped and val:
                    mapped[db_field] += '\n' + val
                elif val:
                    mapped[db_field] = val
                elif db_field not in mapped:
                    mapped[db_field] = ''
        if any(mapped.get(f) for f in fields):
            result.append(mapped)
    return result


def _parse_pcba_import_file(file):
    filename = file.filename.lower()
    if not filename.endswith(('.xlsx', '.xls', '.csv')):
        raise ValueError('仅支持 .xlsx、.xls、.csv 格式')
    if filename.endswith('.csv'):
        import csv, io
        stream = io.StringIO(file.read().decode('utf-8-sig'), newline=None)
        reader = csv.reader(stream)
        all_rows = list(reader)
    else:
        try:
            import openpyxl
        except ImportError:
            raise ValueError('请安装openpyxl: pip install openpyxl')
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
        all_rows = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append([str(c).strip() if c is not None else '' for c in row])
        wb.close()
    if len(all_rows) < 2:
        raise ValueError('文件为空或只有表头')
    result = []
    for row in all_rows[1:]:
        def cell(idx):
            if idx >= len(row):
                return ''
            v = row[idx].strip()
            return '' if v == '/' else v
        pcba_code = cell(1)
        if not pcba_code:
            continue
        compat_parts = []
        for idx in [7, 8, 9]:
            v = cell(idx)
            if v:
                compat_parts.append(v)
        result.append({
            'pcba_code': pcba_code,
            'pcb_code': cell(2),
            'pcb_silkscreen': cell(3),
            'latest_version': cell(4),
            'board_name': cell(5),
            'special_note': cell(6),
            'pcba_version_compat': '\n'.join(compat_parts),
            'compat_description': cell(10),
        })
    return result




@hardware_bp.route('/board-compatibility')
@login_required
def board_compatibility_page():
    series = request.args.get('series', 'SMART').upper()
    model = request.args.get('model', '')
    if series not in ('SMART', 'VENUS'):
        series = 'SMART'
    is_admin = session.get('admin_logged_in', False)
    return render_template_string(shared.get_template('BOARD_COMPATIBILITY_HTML'), series=series, model=model, is_admin=is_admin)




